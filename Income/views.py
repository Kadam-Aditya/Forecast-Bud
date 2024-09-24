from asyncio import IncompleteReadError
from django.shortcuts import render, redirect
from .models import Source, UserIncome
from django.core.paginator import Paginator
from UserPreferences.models import UserPreference
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse
import json
import datetime
import calendar
import json
import os
from .models import UserIncome
import json
from django.http import HttpResponse, JsonResponse
from UserPreferences.models import UserPreference
import datetime
import csv
from django.utils.timezone import now
import csv
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from io import BytesIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from django.utils.timezone import now
from io import BytesIO
from openpyxl.utils import get_column_letter 
from django.core.exceptions import ObjectDoesNotExist
# Create your views here.


def search_income(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        income = UserIncome.objects.filter(
            amount__istartswith=search_str, owner=request.user) | UserIncome.objects.filter(
            date__istartswith=search_str, owner=request.user) | UserIncome.objects.filter(
            description__icontains=search_str, owner=request.user) | UserIncome.objects.filter(
            source__icontains=search_str, owner=request.user)
        data = income.values()
        return JsonResponse(list(data), safe=False)


@login_required(login_url='/authentication/login')
def index(request):
    categories = Source.objects.all()
    income = UserIncome.objects.filter(owner=request.user)
    paginator = Paginator(income, 5)
    page_number = request.GET.get('page')
    page_obj = Paginator.get_page(paginator, page_number)
    try:
        user_preference = UserPreference.objects.get(user=request.user)
        currency = user_preference.currency
    except ObjectDoesNotExist:
        currency = "USD"
    context = {
        'income': income,
        'page_obj': page_obj,
        'currency': currency
    }
    return render(request, 'income/index.html', context)


@login_required(login_url='/authentication/login')
def add_income(request):
    sources = Source.objects.all()
    context = {
        'sources': sources,
        'values': request.POST
    }
    if request.method == 'GET':
        return render(request, 'income/add_income.html', context)

    if request.method == 'POST':
        amount = request.POST['amount']

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'income/add_income.html', context)
        description = request.POST['description']
        date = request.POST['income_date']
        source = request.POST['source']

        if not description:
            messages.error(request, 'description is required')
            return render(request, 'income/add_income.html', context)
        
        if not date:
            messages.error(request, 'date is required')
            return render(request, 'income/add_income.html', context)

        UserIncome.objects.create(owner=request.user, amount=amount, date=date,
                                  source=source, description=description)
        messages.success(request, 'Record saved successfully')

        return redirect('income')


@login_required(login_url='/authentication/login')
def income_edit(request, id):
    income = UserIncome.objects.get(pk=id)
    sources = Source.objects.all()
    context = {
        'income': income,
        'values': income,
        'sources': sources
    }
    if request.method == 'GET':
        return render(request, 'income/edit_income.html', context)
    if request.method == 'POST':
        amount = request.POST['amount']

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'income/edit_income.html', context)
        description = request.POST['description']
        date = request.POST['income_date']
        source = request.POST['source']

        if not description:
            messages.error(request, 'description is required')
            return render(request, 'income/edit_income.html', context)
        income.amount = amount
        income. date = date
        income.source = source
        income.description = description

        income.save()
        messages.success(request, 'Record updated  successfully')

        return redirect('income')


def delete_income(request, id):
    income = UserIncome.objects.get(pk=id)
    income.delete()
    messages.success(request, 'record removed')
    return redirect('income')


def income_source_summary(request):
    todays_date = datetime.date.today()
    six_months_ago = todays_date-datetime.timedelta(days=30*6)
    incomes = UserIncome.objects.filter(owner=request.user,
                                      date__gte=six_months_ago, date__lte=todays_date)
    finalrep = {}

    def get_source(income):
        return income.source
    source_list = list(set(map(get_source, incomes)))

    def get_income_source_amount(source):
        amount = 0
        filtered_by_source = incomes.filter(source=source)

        for item in filtered_by_source:
            amount += item.amount
        return amount

    for x in incomes:
        for y in source_list:
            finalrep[y] = get_income_source_amount(y)

    return JsonResponse({'income_source_data': finalrep}, safe=False)

def incomestats_view(request):
    return render(request, 'income/incomestats.html')

def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    timestamp = now().strftime('%Y-%m-%d_%H-%M-%S')  # Generate a timestamp for the filename
    response['Content-Disposition'] = f'attachment; filename=Incomes_{timestamp}.csv'

    writer = csv.writer(response)
    writer.writerow(['Amount', 'Description', 'Source', 'Date'])

    incomes = UserIncome.objects.filter(owner=request.user)

    for income in incomes:
        writer.writerow([income.amount, income.description, income.source, income.date])

    return response

def export_pdf(request):
    # Create a response object with PDF content type
    response = HttpResponse(content_type='application/pdf')
    timestamp = now().strftime('%Y-%m-%d_%H-%M-%S')  # Generate a timestamp for the filename
    response['Content-Disposition'] = f'attachment; filename=Incomes_{timestamp}.pdf'

    # Create a PDF document
    pdf = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Add data to the PDF
    data = [['Amount', 'Description', 'Source', 'Date']]
    incomes = UserIncome.objects.filter(owner=request.user)

    for income in incomes:
        data.append([income.amount, income.description, income.source, income.date])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    pdf.build(elements)

    return response


def export_excel(request):
    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Incomes'

    # Add column headers with bold formatting
    bold_format = Font(bold=True)
    header_row = ['Amount', 'Description', 'Source', 'Date']
    worksheet.append(header_row)

    # Add expense data to the worksheet
    incomes = UserIncome.objects.filter(owner=request.user)
    for income in incomes:
        row_data = [income.amount, income.description, income.source, income.date]
        worksheet.append(row_data)

    # Adjust column width for the 'Date' column
    date_column = worksheet.column_dimensions[get_column_letter(4)]  # Assuming 'Date' is in the 4th column
    date_column.width = 15  # Adjust the width as needed to display the date properly

    # Create a BytesIO buffer to store the Excel file
    output = BytesIO()
    workbook.save(output)

    # Create a response object with Excel content type and file extension
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = now().strftime('%Y-%m-%d_%H-%M-%S')  # Generate a timestamp for the filename
    response['Content-Disposition'] = f'attachment; filename=Incomes_{timestamp}.xlsx'

    return response