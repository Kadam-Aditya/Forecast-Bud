from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Category, Expense
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.paginator import Paginator
import json
from django.http import HttpResponse, JsonResponse
from UserPreferences.models import UserPreference
import datetime
import csv
from django.utils.timezone import now
import csv
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from io import BytesIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from django.utils.timezone import now
from io import BytesIO
from openpyxl.utils import get_column_letter 
from django.core.exceptions import ObjectDoesNotExist



def search_expenses(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        expenses = Expense.objects.filter(
            amount__istartswith=search_str, owner=request.user) | Expense.objects.filter(
            date__istartswith=search_str, owner=request.user) | Expense.objects.filter(
            description__icontains=search_str, owner=request.user) | Expense.objects.filter(
            category__icontains=search_str, owner=request.user)
        data = expenses.values()
        return JsonResponse(list(data), safe=False)

@login_required(login_url='/authentication/login')
def index(request):
    categories = Category.objects.all()
    expenses = Expense.objects.filter(owner=request.user).order_by('-date') 
    paginator = Paginator(expenses, 6)
    page_number = request.GET.get('page')
    page_obj = Paginator.get_page(paginator, page_number)
    try:
        user_preference = UserPreference.objects.get(user=request.user)
        currency = user_preference.currency
    except ObjectDoesNotExist:
        currency = "USD"
    context = {
        'expenses': expenses,
        'page_obj': page_obj,
        'currency': currency,
    }
    return render(request,'expenses/index.html', context)

@login_required(login_url='/authentication/login')
def add_expense(request):
    categories = Category.objects.all()
    context = {
        'categories': categories,
        'values': request.POST
    }
    if request.method == 'GET':
        return render(request, 'expenses/add_expense.html', context)

    if request.method == 'POST':
        amount = request.POST['amount']

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'expenses/add_expense.html', context)
        description = request.POST['description']
        date = request.POST['expense_date']
        category = request.POST['category']

        if not description:
            messages.error(request, 'description is required')
            return render(request, 'expenses/add_expense.html', context)
        
        if not date:
            messages.error(request, 'Date is required')
            return render(request, 'expenses/add_expense.html', context)
        category = request.POST['category']

        Expense.objects.create(owner=request.user, amount=amount, date=date,
                               category=category, description=description)
        messages.success(request, 'Expense saved successfully')

        return redirect('expenses')
    


@login_required(login_url='/authentication/login')
def expense_edit(request, id):
    expense = Expense.objects.get(pk=id)
    categories = Category.objects.all()
    context = {
        'expense': expense,
        'values': expense,
        'categories': categories
    }
    if request.method == 'GET':
        return render(request, 'expenses/edit-expense.html', context)
    if request.method == 'POST':
        amount = request.POST['amount']

        if not amount:
            messages.error(request, 'Amount is required')
            return render(request, 'expenses/edit-expense.html', context)
        description = request.POST['description']
        date = request.POST['expense_date']
        category = request.POST['category']

        if not description:
            messages.error(request, 'description is required')
            return render(request, 'expenses/edit-expense.html', context)

        expense.owner = request.user
        expense.amount = amount
        expense.date = date
        expense.category = category
        expense.description = description

        expense.save()
        messages.success(request, 'Expense updated  successfully')

        return redirect('expenses')
    
def delete_expense(request, id):
    expense = Expense.objects.get(pk=id)
    expense.delete()
    messages.success(request, 'Expense removed')
    return redirect('expenses')


def expense_category_summary(request):
    todays_date = datetime.date.today()
    six_months_ago = todays_date-datetime.timedelta(days=30*6)
    expenses = Expense.objects.filter(owner=request.user,
                                      date__gte=six_months_ago, date__lte=todays_date)
    finalrep = {}

    def get_category(expense):
        return expense.category
    category_list = list(set(map(get_category, expenses)))

    def get_expense_category_amount(category):
        amount = 0
        filtered_by_category = expenses.filter(category=category)

        for item in filtered_by_category:
            amount += item.amount
        return amount

    for x in expenses:
        for y in category_list:
            finalrep[y] = get_expense_category_amount(y)

    return JsonResponse({'expense_category_data': finalrep}, safe=False)


def stats_view(request):
    return render(request, 'expenses/stats.html')

def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    timestamp = now().strftime('%Y-%m-%d_%H-%M-%S')  # Generate a timestamp for the filename
    response['Content-Disposition'] = f'attachment; filename=Expenses_{timestamp}.csv'

    writer = csv.writer(response)
    writer.writerow(['Amount', 'Description', 'Category', 'Date'])

    expenses = Expense.objects.filter(owner=request.user)

    for expense in expenses:
        writer.writerow([expense.amount, expense.description, expense.category, expense.date])

    return response

def export_pdf(request):
    # Create a response object with PDF content type
    response = HttpResponse(content_type='application/pdf')
    timestamp = now().strftime('%Y-%m-%d_%H-%M-%S')  # Generate a timestamp for the filename
    response['Content-Disposition'] = f'attachment; filename=Expenses_{timestamp}.pdf'

    # Create a PDF document
    pdf = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Add data to the PDF
    data = [['Amount', 'Description', 'Category', 'Date']]
    expenses = Expense.objects.filter(owner=request.user)

    for expense in expenses:
        data.append([expense.amount, expense.description, expense.category, expense.date])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    pdf.build(elements)

    return response


def export_excel(request):
    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Expenses'

    # Add column headers with bold formatting
    bold_format = Font(bold=True)
    header_row = ['Amount', 'Description', 'Category', 'Date']
    worksheet.append(header_row)

    # Add expense data to the worksheet
    expenses = Expense.objects.filter(owner=request.user)
    for expense in expenses:
        row_data = [expense.amount, expense.description, expense.category, expense.date]
        worksheet.append(row_data)

    # Adjust column width for the 'Date' column
    date_column = worksheet.column_dimensions[get_column_letter(4)]  # Assuming 'Date' is in the 4th column
    date_column.width = 15  # Adjust the width as needed to display the date properly

    # Create a BytesIO buffer to store the Excel file
    output = BytesIO()
    workbook.save(output)

    # Create a response object with Excel content type and file extension
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = now().strftime('%Y-%m-%d_%H-%M-%S')  # Generate a timestamp for the filename
    response['Content-Disposition'] = f'attachment; filename=Expenses_{timestamp}.xlsx'

    return response